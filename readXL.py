# Description: This file reads the excel file and returns the data in a dictionary
import openpyxl
from fetch import fetch_url
from deleteXL import delete_XL
from decimal import Decimal, getcontext

def read_neft_XL():
    '''
    This function reads the excel file and returns the data in a dictionary
    '''
    # Define variable to load the dataframe
    dataframe = openpyxl.load_workbook("test.xlsx")

    dataframe1 = dataframe["NEFT"]

    # Iterate the loop to read the cell values
    sheet_data = dict()
    getcontext().prec = 2
    count = 1
    for row in range(4, dataframe1.max_row - 2):
        data = list()
        data.append("2023")
        data.append("JAN")
        for col in dataframe1.iter_cols(2, dataframe1.max_column):
            #print(col[row].value, end=" : ")
            if type(col[row].value) == float:
                data.append(Decimal(col[row].value))
            else:
                data.append(col[row].value)
        #print(" ")
        sheet_data[str(count)] = tuple(data)
        count += 1
    return sheet_data

def read_rtgs_XL():
    '''
    This function reads the excel file and returns the data in a dictionary
    '''
    # Define variable to load the dataframe
    dataframe = openpyxl.load_workbook("test.xlsx")

    dataframe1 = dataframe["RTGS"]

    # Iterate the loop to read the cell values
    sheet_data = dict()
    getcontext().prec = 2
    count = 1
    for row in range(5, dataframe1.max_row - 2):
        data = list()
        data.append("2023")
        data.append("JAN")
        for col in dataframe1.iter_cols(2, dataframe1.max_column-1):
            #print(col[row].value, end=" : ")
            if type(col[row].value) == float:
                data.append(Decimal(col[row].value))
            else:
                data.append(col[row].value)
        #print(" ")
        sheet_data[str(count)] = tuple(data)
        count += 1
    return sheet_data
if __name__ == '__main__':
    fetch_url("https://rbidocs.rbi.org.in/rdocs/NEFT/DOCs/NEFTRTGSJAN20238F691C3E20284DB9827EE56F372B87C0.XLSX")
    x = read_rtgs_XL()
    print(x["2"])
    delete_XL()